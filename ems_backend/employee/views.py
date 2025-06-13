from django.shortcuts import get_object_or_404
from rest_framework import generics
from .models import Employee
from .serializers import EmployeeSerializer
from accounts.permissions import IsAdmin
from rest_framework.permissions import IsAuthenticated
from accounts.permissions import IsSelfOrAdmin
from rest_framework.exceptions import PermissionDenied
from rest_framework.views import APIView
from rest_framework.response import Response



class EmployeeListCreateView(generics.ListCreateAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return Employee.objects.all()
        return Employee.objects.none()  



class EmployeeDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        user = self.request.user

        if user.role == 'admin':
            # Admin can access employee by ID in URL
            return super().get_object()

        elif user.role == 'employee':
            # Ignore the URL ID and always return the logged-in employee's own data
            try:
                return Employee.objects.get(user=user)
            except Employee.DoesNotExist:
                raise PermissionDenied("You do not have permission to access this resource.")

        raise PermissionDenied("Invalid role.")



class EmployeeMeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'employee':
            return Response({'detail': 'Only employees can access this.'}, status=403)

        employee = get_object_or_404(Employee, user=request.user)
        serializer = EmployeeSerializer(employee)
        return Response(serializer.data)



# views.py
import openpyxl
from collections import Counter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .serializers import ExcelUploadSerializer


class ExcelPieChartView(APIView):
    def post(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']

            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active  # Assumes first sheet

                department_list = []
                header = [cell.value for cell in ws[1]]
                dept_index = header.index('Department')

                for row in ws.iter_rows(min_row=2, values_only=True):
                    department = row[dept_index]
                    if department:
                        department_list.append(department)

                dept_counts = Counter(department_list)

                # Format response for pie chart (label, value)
                chart_data = [
                    {"label": dept, "value": count}
                    for dept, count in dept_counts.items()
                ]

                return Response(chart_data, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
